from openpyxl import load_workbook

wb = load_workbook(filename=r'5_21_20180920_mac.xlsx')
ws_mac = wb["5_21"]
wb1 = load_workbook(filename=r'2018-09-25_sn.xlsx')
ws_sn = wb1['Sheet2']


for i in range(500, 6017):
    mac = ws_mac.cell(row=i, column=4).value
    ws_mac.cell(row=i, column=4).value = mac.replace(":","")
    for j in range(2, 138894):
        mac2 = ws_sn.cell(row=j, column=2).value
        if mac == mac2:
            ws_mac.cell(row=i, column=5).value = ws_sn.cell(row=j, column=1).value
            ws_mac.cell(row=i, column=6).value = ws_sn.cell(row=j, column=3).value

wb.save(r'5_21_20180920_mac.xlsx')
