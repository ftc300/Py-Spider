import requests
import json

import time
from openpyxl import Workbook
from openpyxl import load_workbook


def requestDeviceInfo(did):
    url = 'https://iot.mi.com/if/product_device_info?region=local&did=' + did
    headers = {'Accept': 'application/json, text/plain, */*',
               'Referer': 'https://iot.mi.com/deviceInfo.html',
               'Host': 'iot.mi.com',
               'Connection': 'keep-alive',
               'Accept-Encoding': 'gzip, deflate, br',
               'user-agent': 'Mozilla/5.0 (Windows NT 6.1; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/68.0.3440.106 Safari/537.36',
               'Cookie': 'developerStatus=3; userId=1208627857; serviceToken=rhF7gMsTGE0x6/uJJfd0pNiVXFsdxSosoIRqCdxq/oiy9lcxBHl2IseXlchSQA5gMpF0a8yVS5T8gklGJRyK7JvKwNTp78t1isJ6vfKysFKqKWOP5t7jwbpZRQryZAhMpEBInImrWdlaF83IvkSqDXwz8f1vnvXTn6jfzj2eGNaM+5HN/JTlAojgT+fRwVBzXRmNSfBZa6iE5JaVPJPFxgPnQaKWDAvH/PzB1VkbGdw=; cUserId=3GaTuC1IgAFIrCD1peAlqDJHLTU; userName=%E9%9A%90%E7%A7%80%E7%A7%91%E6%8A%80; userIcon=https%3A%2F%2Fs1.mi-img.com%2Fmfsv2%2Favatar%2Ffdsc3%2Fp01ksikvgswH%2F5nw9VSvcd0jJDL.jpg'
               }
    # cookies = dict(axmuid='nMya4uNOI4blUDnTS94ivGD%2FErmarBmM0LzrA9wKQaM%3D', cUserId='3GaTuC1IgAFIrCD1peAlqDJHLTU',
    #                developerStatus='3', Hm_lvt_4982d57ea12df95a2b24715fb6440726='1535710925',
    #                mstuid='1535016408297_4374',
    #                serviceToken='rhF7gMsTGE0x6/uJJfd0pNiVXFsdxSosoIRqCdxq/oiy9lcxBHl2IseXlchSQA5gHgaZcig77YyfWgCMk67shUrqZ7Wsb2Oj2pzsgN0G5tbkyy2RWern8pW90/HTuzScaOw/zX83n/e5T12/7uU97wS4pHGU1pH7S4SjxLOM4eYGsqJP1m7SxzYa6MyGsHDbFyVs/7fSWrKc7XJ1V8AznvtyCflMb2cArlCe+bppz6E=',
    #                userIcon='https%3A%2F%2Fs1.mi-img.com%2Fmfsv2%2Favatar%2Ffdsc3%2Fp01ksikvgswH%2F5nw9VSvcd0jJDL.jpg',
    #                userId='1208627857', userName='%E9%9A%90%E7%A7%80%E7%A7%91%E6%8A%80',
    #                xmuuid='XMGUEST-A8C8A2E0-A6B6-11E8-A8D2-AD8224475B0D')
    cookies = dict(developerStatus='3', userId='1208627857',
                   serviceToken='rhF7gMsTGE0x6/uJJfd0pNiVXFsdxSosoIRqCdxq/oiy9lcxBHl2IseXlchSQA5gMpF0a8yVS5T8gklGJRyK7JvKwNTp78t1isJ6vfKysFKqKWOP5t7jwbpZRQryZAhMpEBInImrWdlaF83IvkSqDXwz8f1vnvXTn6jfzj2eGNaM+5HN/JTlAojgT+fRwVBzXRmNSfBZa6iE5JaVPJPFxgPnQaKWDAvH/PzB1VkbGdw=',
                   cUserId='3GaTuC1IgAFIrCD1peAlqDJHLTU', userName='%E9%9A%90%E7%A7%80%E7%A7%91%E6%8A%80',
                   userIcon='https%3A%2F%2Fs1.mi-img.com%2Fmfsv2%2Favatar%2Ffdsc3%2Fp01ksikvgswH%2F5nw9VSvcd0jJDL.jpg')

    r = requests.get(url, headers=headers, cookies=cookies)
    return r.text


def getMacByDid(response):
    data2 = json.loads(response)
    return data2['result']['mac']


def getDidList():
    ret = []
    wb = load_workbook(filename=r'data.xlsx')
    ws = wb["5_21"]
    for i in range(2, 6017):
        did = ws.cell(row=i, column=2).value
        d = {"row": i, "did": did}
        ret.append(dict(d))
    return ret


if __name__ == '__main__':
    listOfMac = []
    listOfDid = getDidList()
    wb = load_workbook(filename=r'data.xlsx')
    for item in listOfDid:
        try:
            rowId = item["row"]
            did = item["did"]
            response = requestDeviceInfo(did)
            mac = getMacByDid(response)
            # if rowId % 10 == 0:
            #     time.sleep(2)
            # else:
            # time.sleep(0.1)
        except:
            print("rowId:"+ str(rowId) + "throw exception")
            continue
        else:
            d = {"row": rowId, "mac": mac}
            print(d)
            listOfMac.append(dict(d))

    for item in listOfMac:
         ws = wb["5_21"]
         rowId = item["row"]
         mac = item["mac"]
         ws.cell(row=rowId, column=4).value = mac

    wb.save(r'5_21_20180920.xlsx')
