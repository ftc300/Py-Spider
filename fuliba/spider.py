import os
from itertools import chain

from lxml import html
import requests
from openpyxl import Workbook
from openpyxl import load_workbook
from multiprocessing.pool import Pool


def getAllPagesUrl():
    urls = []
    wb = load_workbook(filename=r'data.xlsx')
    wsPage = wb["page"]
    try:
        for i in range(1, 8):
            page = requests.get('http://fulibus.net/category/flhz/page/' + str(i))
            tree = html.fromstring(page.text)
            url = tree.xpath('//div//article//header//h2//@href')
            urls.append(url)
        urls = list(chain(*urls))
        print(urls)
        i = 1
        for url in urls:
            wsPage.cell(row=i, column=1).value = url
            i += 1
    except:
        print("getAllPagesUrl err")
    wb.save(r'data.xlsx')
    return urls


def getFirstPageImgUrls():
    urls = []
    wb = load_workbook(filename=r'data.xlsx')
    wsImg = wb["img"]
    for item in getAllPagesUrl():
        page = requests.get(item)
        tree = html.fromstring(page.text)
        url = tree.xpath('//div//article[@class="article-content"]//p//img//@src')
        print(url)
        urls.append(url)
    urls = list(chain(*urls))
    i = 1
    for url in urls:
        wsImg.cell(row=i, column=1).value = url
        i += 1
    wb.save(r'data.xlsx')
    return urls


def getSecPageImgUrls():
    urls = []
    wb = load_workbook(filename=r'data.xlsx')
    wsImg = wb["img2"]
    for item in getAllPagesUrl():
        page = requests.get(item + "/2")
        tree = html.fromstring(page.text)
        url = tree.xpath('//div//article[@class="article-content"]//p//img//@src')
        print(url)
        urls.append(url)
    urls = list(chain(*urls))
    i = 1
    for url in urls:
        wsImg.cell(row=i, column=1).value = url
        i += 1
    wb.save(r'data.xlsx')
    return urls



def save_image():
    wb = load_workbook(filename=r'data.xlsx')
    wsImg = wb["img"]
    for i in range(1,9267):
        url = wsImg.cell(row=i, column=1).value
        imageName = url.split("/")[-1]
        print(url)
        print(imageName)
        file_path = "d:\\fuliba\\img\\" + imageName
        try:
            if not os.path.exists(file_path):
                with open(file_path, 'wb') as f:
                    f.write(requests.get(url).content)
                print('Downloaded image path is %s' % file_path)
            else:
                print('Already Downloaded', file_path)
        except:
            print('Failed to Save Image，item %s' % url)
def save_image2():
    wb = load_workbook(filename=r'data.xlsx')
    wsImg = wb["img2"]
    for i in range(1,9267):
        url = wsImg.cell(row=i, column=1).value
        imageName = url.split("/")[-1]
        print(url)
        print(imageName)
        file_path = "d:\\fuliba\\img2\\" + imageName
        try:
            if not os.path.exists(file_path):
                with open(file_path, 'wb') as f:
                    f.write(requests.get(url).content)
                print('Downloaded image path is %s' % file_path)
            else:
                print('Already Downloaded', file_path)
        except:
            print('Failed to Save Image，item %s' % url)


save_image2()
